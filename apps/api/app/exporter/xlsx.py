from typing import List, Dict, Any
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from sqlalchemy.orm import Session

from ..db import SessionLocal
from ..models import Document, Artifact


def format_worksheet_header(ws, title: str):
    """Apply consistent formatting to worksheet headers."""
    # Set title in A1
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Merge cells for title
    ws.merge_cells('A1:F1')
    
    # Add header background
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    for col in range(1, 7):  # A through F
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = Font(color='FFFFFF', bold=True)


def create_index_sheet(workbook: Workbook, doc_id: str, artifacts: List[Artifact]) -> None:
    """Create an index sheet with document summary."""
    ws = workbook.active
    ws.title = "Index"
    
    # Title
    format_worksheet_header(ws, f"Document Export - {doc_id}")
    
    # Document metadata
    row = 3
    ws[f'A{row}'] = "Document ID:"
    ws[f'B{row}'] = doc_id
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Export Date:"
    ws[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Total Artifacts:"
    ws[f'B{row}'] = len(artifacts)
    ws[f'A{row}'].font = Font(bold=True)
    
    # Artifacts summary
    row += 2
    ws[f'A{row}'] = "Artifacts Summary:"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    
    row += 1
    headers = ["Sheet Name", "Type", "Page", "Engine", "Status", "Rows", "Columns"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    
    # Add artifact details
    for i, artifact in enumerate(artifacts):
        row += 1
        sheet_name = f"Table_{i+1}_P{artifact.page}"
        
        # Basic info
        ws.cell(row=row, column=1, value=sheet_name)
        ws.cell(row=row, column=2, value=artifact.kind.title())
        ws.cell(row=row, column=3, value=artifact.page)
        ws.cell(row=row, column=4, value=artifact.engine.title())
        ws.cell(row=row, column=5, value=artifact.status.title())
        
        # Table dimensions
        payload = artifact.payload
        if isinstance(payload, dict) and 'rows' in payload:
            rows_data = payload['rows']
            if rows_data:
                ws.cell(row=row, column=6, value=len(rows_data))
                ws.cell(row=row, column=7, value=len(rows_data[0]) if rows_data[0] else 0)
            else:
                ws.cell(row=row, column=6, value=0)
                ws.cell(row=row, column=7, value=0)
        else:
            ws.cell(row=row, column=6, value="N/A")
            ws.cell(row=row, column=7, value="N/A")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width


def create_table_sheet(workbook: Workbook, artifact: Artifact, sheet_index: int) -> None:
    """Create a worksheet for a table artifact."""
    sheet_name = f"Table_{sheet_index}_P{artifact.page}"
    ws = workbook.create_sheet(title=sheet_name)
    
    # Title
    format_worksheet_header(ws, f"Table from Page {artifact.page}")
    
    # Provenance information
    row = 3
    provenance_data = [
        ("Document ID:", artifact.document_id),
        ("Page:", artifact.page),
        ("Extraction Engine:", artifact.engine.title()),
        ("Extraction Time:", artifact.created_at.strftime("%Y-%m-%d %H:%M:%S") if artifact.created_at else "Unknown"),
        ("Status:", artifact.status.title()),
    ]
    
    for label, value in provenance_data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # Table data
    payload = artifact.payload
    if isinstance(payload, dict) and 'rows' in payload:
        rows_data = payload['rows']
        
        if rows_data:
            row += 1
            ws[f'A{row}'] = "Table Data:"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            # Add table headers and data
            for i, table_row in enumerate(rows_data):
                for j, cell_value in enumerate(table_row):
                    cell = ws.cell(row=row + i, column=j + 1, value=cell_value)
                    
                    # Format header row
                    if i == 0:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
            
            # Quality checks
            row += len(rows_data) + 2
            ws[f'A{row}'] = "Quality Checks:"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            # Calculate basic statistics
            total_cells = len(rows_data) * len(rows_data[0]) if rows_data and rows_data[0] else 0
            empty_cells = 0
            numeric_cells = 0
            
            for table_row in rows_data[1:]:  # Skip header
                for cell_value in table_row:
                    if not cell_value or str(cell_value).strip() == "":
                        empty_cells += 1
                    else:
                        # Check if numeric
                        try:
                            float(str(cell_value).replace(',', '').replace('$', '').replace('%', ''))
                            numeric_cells += 1
                        except (ValueError, TypeError):
                            pass
            
            data_cells = total_cells - len(rows_data[0]) if rows_data else 0  # Exclude header
            blank_rate = (empty_cells / max(1, data_cells)) * 100
            numeric_rate = (numeric_cells / max(1, data_cells)) * 100
            
            checks = [
                ("Total Rows:", len(rows_data)),
                ("Total Columns:", len(rows_data[0]) if rows_data else 0),
                ("Data Cells:", data_cells),
                ("Blank Cell Rate:", f"{blank_rate:.1f}%"),
                ("Numeric Cell Rate:", f"{numeric_rate:.1f}%"),
            ]
            
            for label, value in checks:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
        
        else:
            row += 1
            ws[f'A{row}'] = "No table data available"
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def build_workbook(doc_id: str) -> bytes:
    """Build Excel workbook with document artifacts."""
    db: Session = SessionLocal()
    
    try:
        # Get document and check if it exists
        document = db.query(Document).filter(Document.id == doc_id).first()
        if not document:
            raise ValueError(f"Document {doc_id} not found")
        
        # Get all table artifacts for this document
        artifacts = db.query(Artifact).filter(
            Artifact.document_id == doc_id,
            Artifact.kind == "table"
        ).order_by(Artifact.page, Artifact.created_at).all()
        
        # Create workbook
        workbook = Workbook()
        
        # Create index sheet
        create_index_sheet(workbook, doc_id, artifacts)
        
        # Create sheet for each table artifact
        for i, artifact in enumerate(artifacts, 1):
            create_table_sheet(workbook, artifact, i)
        
        # If no table artifacts, add a note
        if not artifacts:
            ws = workbook.create_sheet(title="No_Tables")
            ws['A1'] = "No table artifacts found for this document"
            ws['A1'].font = Font(bold=True, size=12)
        
        # Save to bytes
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    finally:
        db.close()