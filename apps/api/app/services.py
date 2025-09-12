import json
import logging
import time
import uuid
from dataclasses import dataclass
from typing import Optional, List, Dict, Any, Union, Generic, TypeVar
from sqlalchemy.orm import Session, joinedload
from sqlalchemy.exc import IntegrityError, SQLAlchemyError
from sqlalchemy import func
from .models import Document, ProcessingEvent, ProcessingStatus, EventType, Artifact
from .aws import s3_manager
from .db import db_manager
from .metrics import get_metrics_collector
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

logger = logging.getLogger(__name__)

T = TypeVar('T')

@dataclass
class ServiceResult(Generic[T]):
    """Standardized result wrapper for service operations."""
    success: bool
    data: Optional[T] = None
    error: Optional[str] = None
    error_code: Optional[str] = None
    metadata: Optional[Dict[str, Any]] = None
    
    @classmethod
    def success_result(cls, data: T, metadata: Optional[Dict[str, Any]] = None) -> 'ServiceResult[T]':
        return cls(success=True, data=data, metadata=metadata)
    
    @classmethod
    def error_result(cls, error: str, error_code: Optional[str] = None, 
                    metadata: Optional[Dict[str, Any]] = None) -> 'ServiceResult[T]':
        return cls(success=False, error=error, error_code=error_code, metadata=metadata)

class DocumentService:
    """Enhanced document service with proper error handling and retry logic."""
    
    @staticmethod
    def create_document(
        s3_key: str, 
        original_filename: str, 
        content_type: str, 
        file_size: int, 
        sha256_hash: Optional[str] = None
    ) -> ServiceResult[Document]:
        """Create a new document with comprehensive error handling and audit trail."""
        start_time = time.time()
        
        try:
            with db_manager.get_session() as session:
                # Check for duplicate by S3 key or hash
                existing_doc = None
                if sha256_hash:
                    existing_doc = session.query(Document).filter(
                        Document.sha256_hash == sha256_hash
                    ).first()
                
                if not existing_doc:
                    existing_doc = session.query(Document).filter(
                        Document.s3_key == s3_key
                    ).first()
                
                if existing_doc:
                    logger.warning(f"Document already exists: {existing_doc.id}")
                    return ServiceResult.error_result(
                        "Document already exists",
                        "DUPLICATE_DOCUMENT",
                        {"existing_id": existing_doc.id}
                    )
                
                # Create new document
                doc = Document(
                    id=str(uuid.uuid4()),
                    s3_key=s3_key,
                    original_filename=original_filename,
                    content_type=content_type,
                    file_size=file_size,
                    sha256_hash=sha256_hash,
                    processing_status=ProcessingStatus.UPLOADED
                )
                session.add(doc)
                
                # Log creation event
                event = ProcessingEvent(
                    document_id=doc.id,
                    event_type=EventType.DOCUMENT_UPLOADED,
                    message=f"Document uploaded: {original_filename}",
                    event_metadata=json.dumps({
                        "file_size": file_size, 
                        "content_type": content_type,
                        "sha256_hash": sha256_hash,
                        "creation_time_ms": round((time.time() - start_time) * 1000, 2)
                    })
                )
                session.add(event)
                
                # Commit transaction
                session.commit()
                
                processing_time = time.time() - start_time
                logger.info(f"Document created successfully: {doc.id} in {processing_time:.3f}s")
                
                # Record metrics
                metrics = get_metrics_collector()
                metrics.record_document_upload("success")
                metrics.record_database_query("create", "documents", processing_time)
                
                return ServiceResult.success_result(
                    doc, 
                    {"processing_time_ms": round(processing_time * 1000, 2)}
                )
                
        except IntegrityError as e:
            logger.error(f"Document creation integrity error: {e}")
            return ServiceResult.error_result(
                "Document already exists or constraint violation",
                "INTEGRITY_ERROR"
            )
        except SQLAlchemyError as e:
            logger.error(f"Database error during document creation: {e}")
            return ServiceResult.error_result(
                "Database operation failed",
                "DATABASE_ERROR"
            )
        except Exception as e:
            logger.error(f"Unexpected error during document creation: {e}")
            return ServiceResult.error_result(
                "Internal server error",
                "INTERNAL_ERROR"
            )

    @staticmethod
    def get_document(doc_id: str) -> ServiceResult[Optional[Document]]:
        """Get document by ID with proper error handling and eager loading."""
        start_time = time.time()
        
        try:
            with db_manager.get_session() as session:
                doc = session.query(Document).options(
                    joinedload(Document.pages),
                    joinedload(Document.events),
                    joinedload(Document.artifacts)
                ).filter(Document.id == doc_id).first()
                
                processing_time = time.time() - start_time
                
                if doc:
                    logger.debug(f"Document retrieved: {doc_id} in {processing_time:.3f}s")
                    return ServiceResult.success_result(
                        doc,
                        {"processing_time_ms": round(processing_time * 1000, 2)}
                    )
                else:
                    logger.warning(f"Document not found: {doc_id}")
                    return ServiceResult.error_result(
                        "Document not found",
                        "NOT_FOUND",
                        {"document_id": doc_id}
                    )
                    
        except SQLAlchemyError as e:
            logger.error(f"Database error retrieving document {doc_id}: {e}")
            return ServiceResult.error_result(
                "Database operation failed",
                "DATABASE_ERROR"
            )
        except Exception as e:
            logger.error(f"Unexpected error retrieving document {doc_id}: {e}")
            return ServiceResult.error_result(
                "Internal server error",
                "INTERNAL_ERROR"
            )

    @staticmethod
    def get_documents_by_status(status: ProcessingStatus, limit: int = 100) -> ServiceResult[List[Document]]:
        """Get documents by processing status with eager loading."""
        start_time = time.time()
        
        try:
            with db_manager.get_session() as session:
                docs = session.query(Document).options(
                    joinedload(Document.pages),
                    joinedload(Document.events),
                    joinedload(Document.artifacts)
                ).filter(
                    Document.processing_status == status
                ).order_by(Document.created_at.desc()).limit(limit).all()
                
                processing_time = time.time() - start_time
                logger.debug(f"Retrieved {len(docs)} documents with status {status.value} in {processing_time:.3f}s")
                
                return ServiceResult.success_result(
                    docs,
                    {
                        "count": len(docs),
                        "status": status.value,
                        "processing_time_ms": round(processing_time * 1000, 2)
                    }
                )
                
        except SQLAlchemyError as e:
            logger.error(f"Database error retrieving documents by status {status.value}: {e}")
            return ServiceResult.error_result(
                "Database operation failed",
                "DATABASE_ERROR"
            )
        except Exception as e:
            logger.error(f"Unexpected error retrieving documents by status {status.value}: {e}")
            return ServiceResult.error_result(
                "Internal server error",
                "INTERNAL_ERROR"
            )

    @staticmethod
    def update_processing_status(
        doc_id: str, 
        status: ProcessingStatus, 
        error_message: Optional[str] = None,
        additional_metadata: Optional[Dict[str, Any]] = None
    ) -> ServiceResult[Document]:
        """Update document processing status with comprehensive logging."""
        start_time = time.time()
        
        def _perform_update():
            with db_manager.get_session() as session:
                doc = session.query(Document).filter(Document.id == doc_id).first()
                if not doc:
                    return ServiceResult.error_result(
                        f"Document not found: {doc_id}",
                        "NOT_FOUND",
                        {"document_id": doc_id}
                    )
                
                old_status = doc.processing_status
                doc.processing_status = status
                if error_message:
                    doc.error_message = error_message
                
                # Determine event type based on status
                event_type_map = {
                    ProcessingStatus.PROCESSING: EventType.PROCESSING_STARTED,
                    ProcessingStatus.COMPLETED: EventType.PROCESSING_COMPLETED,
                    ProcessingStatus.FAILED: EventType.PROCESSING_FAILED
                }
                event_type = event_type_map.get(status, EventType.PROCESSING_STARTED)
                
                # Create metadata for the event
                event_metadata = {
                    "old_status": old_status.value if old_status else None,
                    "new_status": status.value,
                    "error_message": error_message,
                    "processing_time_ms": round((time.time() - start_time) * 1000, 2)
                }
                
                if additional_metadata:
                    event_metadata.update(additional_metadata)
                
                # Log status change
                event = ProcessingEvent(
                    document_id=doc_id,
                    event_type=event_type,
                    message=f"Status changed from {old_status.value if old_status else 'None'} to {status.value}",
                    event_metadata=json.dumps(event_metadata)
                )
                session.add(event)
                
                session.commit()
                
                processing_time = time.time() - start_time
                logger.info(f"Document {doc_id} status updated to {status.value} in {processing_time:.3f}s")
                
                return ServiceResult.success_result(
                    doc,
                    {"processing_time_ms": round(processing_time * 1000, 2)}
                )
        
        try:
            # Use database retry logic for critical status updates
            return db_manager.retry_on_disconnect(_perform_update, max_retries=3)
            
        except SQLAlchemyError as e:
            logger.error(f"Database error updating document status {doc_id}: {e}")
            return ServiceResult.error_result(
                "Database operation failed",
                "DATABASE_ERROR"
            )
        except Exception as e:
            logger.error(f"Unexpected error updating document status {doc_id}: {e}")
            return ServiceResult.error_result(
                "Internal server error",
                "INTERNAL_ERROR"
            )

    @staticmethod
    def download_document_content(doc_id: str) -> ServiceResult[bytes]:
        """Download document content from S3 with enhanced error handling."""
        start_time = time.time()
        
        try:
            # First get document metadata
            doc_result = DocumentService.get_document(doc_id)
            if not doc_result.success or not doc_result.data:
                return ServiceResult.error_result(
                    doc_result.error or "Document not found",
                    doc_result.error_code or "NOT_FOUND"
                )
            
            doc = doc_result.data
            
            # Download from S3
            try:
                content = s3_manager.download_file(doc.s3_key)
                
                processing_time = time.time() - start_time
                logger.info(f"Document content downloaded: {doc_id} ({len(content)} bytes) in {processing_time:.3f}s")
                
                return ServiceResult.success_result(
                    content,
                    {
                        "file_size": len(content),
                        "s3_key": doc.s3_key,
                        "processing_time_ms": round(processing_time * 1000, 2)
                    }
                )
                
            except Exception as s3_error:
                logger.error(f"S3 download failed for document {doc_id}: {s3_error}")
                
                # Log the failure event
                DocumentService.create_processing_event(
                    doc_id,
                    EventType.PROCESSING_FAILED,
                    f"S3 download failed: {str(s3_error)}",
                    {"s3_key": doc.s3_key, "error": str(s3_error)}
                )
                
                return ServiceResult.error_result(
                    "Failed to download document content",
                    "S3_DOWNLOAD_ERROR",
                    {"s3_key": doc.s3_key}
                )
                
        except Exception as e:
            logger.error(f"Unexpected error downloading document content {doc_id}: {e}")
            return ServiceResult.error_result(
                "Internal server error",
                "INTERNAL_ERROR"
            )

    @staticmethod
    def create_processing_event(
        doc_id: str, 
        event_type: EventType, 
        message: str, 
        metadata: Optional[Dict[str, Any]] = None
    ) -> ServiceResult[ProcessingEvent]:
        """Create a processing event with proper error handling."""
        try:
            with db_manager.get_session() as session:
                event = ProcessingEvent(
                    document_id=doc_id,
                    event_type=event_type,
                    message=message,
                    event_metadata=json.dumps(metadata or {})
                )
                session.add(event)
                session.commit()
                
                logger.debug(f"Processing event created for document {doc_id}: {event_type.value}")
                return ServiceResult.success_result(event)
                
        except SQLAlchemyError as e:
            logger.error(f"Database error creating processing event for {doc_id}: {e}")
            return ServiceResult.error_result(
                "Failed to create processing event",
                "DATABASE_ERROR"
            )
        except Exception as e:
            logger.error(f"Unexpected error creating processing event for {doc_id}: {e}")
            return ServiceResult.error_result(
                "Internal server error",
                "INTERNAL_ERROR"
            )

    @staticmethod
    def get_processing_events(doc_id: str, limit: int = 50) -> ServiceResult[List[ProcessingEvent]]:
        """Get processing events for a document."""
        try:
            with db_manager.get_session() as session:
                events = session.query(ProcessingEvent).filter(
                    ProcessingEvent.document_id == doc_id
                ).order_by(ProcessingEvent.created_at.desc()).limit(limit).all()
                
                return ServiceResult.success_result(
                    events,
                    {"count": len(events), "document_id": doc_id}
                )
                
        except SQLAlchemyError as e:
            logger.error(f"Database error retrieving processing events for {doc_id}: {e}")
            return ServiceResult.error_result(
                "Database operation failed",
                "DATABASE_ERROR"
            )
        except Exception as e:
            logger.error(f"Unexpected error retrieving processing events for {doc_id}: {e}")
            return ServiceResult.error_result(
                "Internal server error",
                "INTERNAL_ERROR"
            )

    @staticmethod
    def delete_document(doc_id: str, delete_from_s3: bool = True) -> ServiceResult[bool]:
        """Delete document and optionally its S3 content."""
        start_time = time.time()
        
        try:
            with db_manager.get_session() as session:
                doc = session.query(Document).filter(Document.id == doc_id).first()
                if not doc:
                    return ServiceResult.error_result(
                        f"Document not found: {doc_id}",
                        "NOT_FOUND"
                    )
                
                s3_key = doc.s3_key
                
                # Delete from S3 if requested
                if delete_from_s3:
                    try:
                        s3_manager.delete_file(s3_key)
                        logger.info(f"S3 file deleted: {s3_key}")
                    except Exception as s3_error:
                        logger.error(f"Failed to delete S3 file {s3_key}: {s3_error}")
                        # Continue with database deletion even if S3 deletion fails
                
                # Delete processing events first (foreign key constraint)
                session.query(ProcessingEvent).filter(
                    ProcessingEvent.document_id == doc_id
                ).delete()
                
                # Delete document
                session.delete(doc)
                session.commit()
                
                processing_time = time.time() - start_time
                logger.info(f"Document deleted: {doc_id} in {processing_time:.3f}s")
                
                return ServiceResult.success_result(
                    True,
                    {
                        "document_id": doc_id,
                        "s3_key": s3_key,
                        "deleted_from_s3": delete_from_s3,
                        "processing_time_ms": round(processing_time * 1000, 2)
                    }
                )
                
        except SQLAlchemyError as e:
            logger.error(f"Database error deleting document {doc_id}: {e}")
            return ServiceResult.error_result(
                "Database operation failed",
                "DATABASE_ERROR"
            )
        except Exception as e:
            logger.error(f"Unexpected error deleting document {doc_id}: {e}")
            return ServiceResult.error_result(
                "Internal server error",
                "INTERNAL_ERROR"
            )

    @staticmethod
    def get_document_stats() -> ServiceResult[Dict[str, Any]]:
        """Get document statistics with optimized queries."""
        try:
            with db_manager.get_session() as session:
                # Use a single query to get all statistics
                stats_query = session.query(
                    func.count(Document.id).label('total_documents'),
                    func.sum(Document.file_size).label('total_size'),
                    func.avg(Document.file_size).label('avg_size'),
                    func.count(Document.id).filter(Document.processing_status == ProcessingStatus.UPLOADED).label('uploaded_count'),
                    func.count(Document.id).filter(Document.processing_status == ProcessingStatus.PROCESSING).label('processing_count'),
                    func.count(Document.id).filter(Document.processing_status == ProcessingStatus.COMPLETED).label('completed_count'),
                    func.count(Document.id).filter(Document.processing_status == ProcessingStatus.FAILED).label('failed_count'),
                    func.count(Document.id).filter(Document.processing_status == ProcessingStatus.RETRYING).label('retrying_count')
                ).first()
                
                if stats_query:
                    stats = {
                        "total_documents": stats_query.total_documents or 0,
                        "status_counts": {
                            "uploaded": stats_query.uploaded_count or 0,
                            "processing": stats_query.processing_count or 0,
                            "completed": stats_query.completed_count or 0,
                            "failed": stats_query.failed_count or 0,
                            "retrying": stats_query.retrying_count or 0
                        },
                        "total_file_size_bytes": stats_query.total_size or 0,
                        "average_file_size_bytes": round(stats_query.avg_size or 0, 2),
                        "timestamp": time.time()
                    }
                else:
                    stats = {
                        "total_documents": 0,
                        "status_counts": {status.value: 0 for status in ProcessingStatus},
                        "total_file_size_bytes": 0,
                        "average_file_size_bytes": 0,
                        "timestamp": time.time()
                    }
                
                return ServiceResult.success_result(stats)
                
        except SQLAlchemyError as e:
            logger.error(f"Database error getting document stats: {e}")
            return ServiceResult.error_result(
                "Database operation failed",
                "DATABASE_ERROR"
            )
        except Exception as e:
            logger.error(f"Unexpected error getting document stats: {e}")
            return ServiceResult.error_result(
                "Internal server error",
                "INTERNAL_ERROR"
            )

    @staticmethod
    def get_extraction_artifacts(doc_id: str) -> ServiceResult[List[Artifact]]:
        """Get all extraction artifacts for a document."""
        try:
            with db_manager.get_session() as session:
                artifacts = session.query(Artifact).filter(
                    Artifact.document_id == doc_id,
                    Artifact.artifact_type == 'table'
                ).order_by(Artifact.page_id, Artifact.id).all()
                
                return ServiceResult.success_result(
                    artifacts,
                    {"count": len(artifacts), "document_id": doc_id}
                )
                
        except SQLAlchemyError as e:
            logger.error(f"Database error retrieving artifacts for {doc_id}: {e}")
            return ServiceResult.error_result(
                "Database operation failed",
                "DATABASE_ERROR"
            )
        except Exception as e:
            logger.error(f"Unexpected error retrieving artifacts for {doc_id}: {e}")
            return ServiceResult.error_result(
                "Internal server error",
                "INTERNAL_ERROR"
            )

    @staticmethod
    def generate_excel_output(document_id: str) -> ServiceResult[bytes]:
        """Generate Excel file with extracted table data."""
        start_time = time.time()
        
        try:
            # Get document and artifacts
            doc_result = DocumentService.get_document(document_id)
            if not doc_result.success or not doc_result.data:
                return ServiceResult.error_result(
                    "Document not found",
                    "NOT_FOUND",
                    {"document_id": document_id}
                )
            
            artifacts_result = DocumentService.get_extraction_artifacts(document_id)
            if not artifacts_result.success:
                return ServiceResult.error_result(
                    "Failed to retrieve extraction artifacts",
                    "ARTIFACTS_ERROR"
                )
            
            artifacts = artifacts_result.data
            if not artifacts:
                return ServiceResult.error_result(
                    "No extraction artifacts found",
                    "NO_ARTIFACTS",
                    {"document_id": document_id}
                )
            
            # Generate Excel file
            excel_bytes = _create_excel_from_artifacts(artifacts, document_id)
            
            processing_time = time.time() - start_time
            logger.info(f"Generated Excel for {document_id} in {processing_time:.3f}s")
            
            # Record metrics
            metrics = get_metrics_collector()
            metrics.record_excel_export("success", processing_time)
            
            return ServiceResult.success_result(
                excel_bytes,
                {
                    "file_size": len(excel_bytes),
                    "artifacts_count": len(artifacts),
                    "processing_time_ms": round(processing_time * 1000, 2)
                }
            )
            
        except Exception as e:
            processing_time = time.time() - start_time
            logger.error(f"Excel generation failed for {document_id}: {e}")
            return ServiceResult.error_result(
                "Excel generation failed",
                "EXCEL_ERROR",
                {"error": str(e), "processing_time_ms": round(processing_time * 1000, 2)}
            )

def _create_excel_from_artifacts(artifacts: List[Artifact], document_id: str) -> bytes:
    """Create Excel file from extraction artifacts."""
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    sheet_count = 0
    
    for artifact in artifacts:
        try:
            # Parse artifact data
            if not artifact.data:
                continue
                
            table_data = json.loads(artifact.data)
            if not table_data.get('data'):
                continue
            
            # Create DataFrame
            df = pd.DataFrame(table_data['data'])
            if df.empty:
                continue
            
            # Apply ledger transformations
            df = _apply_ledger_transformations(df)
            
            # Create worksheet
            page_num = table_data.get('page', 1)
            engine = table_data.get('engine', 'unknown')
            sheet_name = f"Page_{page_num}_{engine}"
            
            # Ensure sheet name is valid (max 31 chars, no special chars)
            sheet_name = _sanitize_sheet_name(sheet_name)
            if len(sheet_name) > 31:
                sheet_name = f"Page_{page_num}_{sheet_count + 1}"
            
            ws = wb.create_sheet(title=sheet_name)
            
            # Add data to worksheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Style the header row
            if ws.max_row > 0:
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add metadata as comments
            ws.cell(row=1, column=1).comment = f"Extracted by: {engine}\nConfidence: {table_data.get('accuracy', 'N/A')}\nRows: {len(df)}, Cols: {len(df.columns)}"
            
            sheet_count += 1
            
        except Exception as e:
            logger.warning(f"Failed to process artifact {artifact.id}: {e}")
            continue
    
    # If no sheets were created, create a summary sheet
    if sheet_count == 0:
        ws = wb.create_sheet(title="Summary")
        ws.append(["No tables extracted from this document"])
        ws.append(["Please check if the document contains tabular data"])
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    return output.getvalue()

def _sanitize_sheet_name(name: str) -> str:
    """Sanitize sheet name for Excel compatibility."""
    # Remove invalid characters
    invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
    for char in invalid_chars:
        name = name.replace(char, '_')
    
    # Limit length
    if len(name) > 31:
        name = name[:31]
    
    return name

def _apply_ledger_transformations(df: pd.DataFrame) -> pd.DataFrame:
    """Apply business logic transformations specific to ledger data."""
    if df.empty:
        return df
    
    # Create a copy to avoid modifying original
    transformed_df = df.copy()
    
    try:
        # Common ledger transformations
        for col in transformed_df.columns:
            # Convert numeric columns
            if transformed_df[col].dtype == 'object':
                # Try to convert to numeric, keeping original if fails
                numeric_series = pd.to_numeric(transformed_df[col], errors='coerce')
                if not numeric_series.isna().all():
                    transformed_df[col] = numeric_series
        
        # Remove completely empty rows
        transformed_df = transformed_df.dropna(how='all')
        
        # Remove completely empty columns
        transformed_df = transformed_df.dropna(axis=1, how='all')
        
        # Standardize column names (remove extra whitespace, etc.)
        transformed_df.columns = [str(col).strip() for col in transformed_df.columns]
        
        logger.debug(f"Applied ledger transformations: {len(transformed_df)} rows, {len(transformed_df.columns)} columns")
        
    except Exception as e:
        logger.warning(f"Ledger transformations failed: {e}")
    
    return transformed_df